"""
Generate an operational monthly dashboard Excel workbook for business tracking.

Output:
  docs/TABLERO_OPERATIVO_MENSUAL_2026.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT_PATH = Path("docs/TABLERO_OPERATIVO_MENSUAL_2026.xlsx")


def _set_border(cell) -> None:
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _build_instructions_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instrucciones"

    ws["A1"] = "Tablero Operativo Mensual - Instrucciones"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left")

    lines = [
        "",
        "Objetivo: seguimiento mensual de clientes, MRR, conversion y churn.",
        "",
        "Que debes rellenar cada mes (hoja 'Tablero'):",
        "- C: Clientes reales",
        "- F: MRR real (EUR)",
        "- I: Conversion Corrector->Pro real (en porcentaje)",
        "- K: Churn real (en porcentaje)",
        "- L: Horas de soporte/mes",
        "- M: Gasto comercial (EUR/mes)",
        "",
        "Que NO tocar (automatico):",
        "- D y G: deltas vs objetivo",
        "- N: semaforo (VERDE/AMARILLO/ROJO)",
        "- O: accion recomendada",
        "",
        "Reglas de semaforo:",
        "- VERDE: clientes y MRR >= 95% del objetivo y churn dentro de objetivo",
        "- AMARILLO: clientes o MRR entre 85% y 94% o churn > objetivo",
        "- ROJO: clientes o MRR < 85% o churn > objetivo + 1 punto",
        "",
        "Que hacer segun resultado:",
        "- VERDE: mantener plan, optimizar conversion y referral",
        "- AMARILLO: ajuste tactico (onboarding, mensaje de valor, canal comercial)",
        "- ROJO: plan de choque 30 dias (retencion+producto+captacion)",
        "",
        "Consejo operativo:",
        "- Cierre semanal: actualizar reales del mes en curso",
        "- Cierre mensual: registrar decisiones y responsables para el mes siguiente",
    ]

    row = 2
    for line in lines:
        ws[f"A{row}"] = line
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.column_dimensions["A"].width = 120


def _build_dashboard_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Tablero")

    ws["A1"] = "Tablero Operativo Mensual 2026"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "ARPU neto referencia (editable)"
    ws["B2"] = 35.62
    ws["B2"].number_format = "#,##0.00"

    ws["D2"] = "Objetivo anual clientes"
    ws["E2"] = 90

    headers = [
        "Mes",
        "Clientes_obj",
        "Clientes_real",
        "Delta_clientes_pct",
        "MRR_obj_EUR",
        "MRR_real_EUR",
        "Delta_MRR_pct",
        "Conv_C2P_obj",
        "Conv_C2P_real",
        "Churn_obj",
        "Churn_real",
        "Soporte_h",
        "Comercial_EUR",
        "Semaforo",
        "Accion_recomendada",
    ]

    header_row = 5
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _set_border(cell)

    months = [f"M{i}" for i in range(1, 13)]
    clients_obj = [5, 12, 20, 28, 36, 45, 53, 61, 70, 77, 84, 90]
    conv_obj = [0.005, 0.007, 0.008, 0.010, 0.011, 0.013, 0.014, 0.015, 0.016, 0.017, 0.018, 0.020]
    churn_obj = [0.10, 0.09, 0.08, 0.07, 0.065, 0.06, 0.055, 0.052, 0.05, 0.047, 0.044, 0.04]

    start_row = header_row + 1
    for idx, month in enumerate(months):
        row = start_row + idx

        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=clients_obj[idx])
        ws.cell(row=row, column=4, value=f'=IF(B{row}=0,"",C{row}/B{row}-1)')
        ws.cell(row=row, column=5, value=f"=ROUND(B{row}*$B$2,2)")
        ws.cell(row=row, column=7, value=f'=IF(E{row}=0,"",F{row}/E{row}-1)')
        ws.cell(row=row, column=8, value=conv_obj[idx])
        ws.cell(row=row, column=10, value=churn_obj[idx])
        ws.cell(
            row=row,
            column=14,
            value=(
                f'=IF(OR(C{row}<0.85*B{row},F{row}<0.85*E{row},K{row}>J{row}+0.01),'
                f'"ROJO",IF(OR(C{row}<0.95*B{row},F{row}<0.95*E{row},K{row}>J{row}),'
                f'"AMARILLO","VERDE"))'
            ),
        )
        ws.cell(
            row=row,
            column=15,
            value=(
                f'=IF(N{row}="ROJO","Plan choque 30d: retencion+producto+captacion",'
                f'IF(N{row}="AMARILLO","Ajuste tactico: onboarding+mensaje+canales",'
                f'"Mantener plan y optimizar conversion"))'
            ),
        )

        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            _set_border(cell)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.cell(row=row, column=4).number_format = "0.00%"
        ws.cell(row=row, column=7).number_format = "0.00%"
        ws.cell(row=row, column=8).number_format = "0.00%"
        ws.cell(row=row, column=9).number_format = "0.00%"
        ws.cell(row=row, column=10).number_format = "0.00%"
        ws.cell(row=row, column=11).number_format = "0.00%"
        ws.cell(row=row, column=5).number_format = "#,##0.00"
        ws.cell(row=row, column=6).number_format = "#,##0.00"
        ws.cell(row=row, column=13).number_format = "#,##0.00"

    red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    range_ref = f"N{start_row}:N{start_row + len(months) - 1}"
    ws.conditional_formatting.add(range_ref, FormulaRule(formula=[f'$N{start_row}="ROJO"'], fill=red_fill))
    ws.conditional_formatting.add(range_ref, FormulaRule(formula=[f'$N{start_row}="AMARILLO"'], fill=yellow_fill))
    ws.conditional_formatting.add(range_ref, FormulaRule(formula=[f'$N{start_row}="VERDE"'], fill=green_fill))

    summary_row = start_row + len(months) + 2
    ws[f"A{summary_row}"] = "Resumen rapido"
    ws[f"A{summary_row}"].font = Font(bold=True)
    ws[f"A{summary_row + 1}"] = "Ultimo mes con datos"
    ws[f"B{summary_row + 1}"] = f'=IFERROR(LOOKUP(2,1/(C{start_row}:C{start_row + 11}<>""),A{start_row}:A{start_row + 11}),"")'
    ws[f"A{summary_row + 2}"] = "Clientes ultimo dato"
    ws[f"B{summary_row + 2}"] = f'=IFERROR(LOOKUP(2,1/(C{start_row}:C{start_row + 11}<>""),C{start_row}:C{start_row + 11}),"")'
    ws[f"A{summary_row + 3}"] = "MRR ultimo dato (EUR)"
    ws[f"B{summary_row + 3}"] = f'=IFERROR(LOOKUP(2,1/(F{start_row}:F{start_row + 11}<>""),F{start_row}:F{start_row + 11}),"")'
    ws[f"A{summary_row + 4}"] = "Churn promedio cargado"
    ws[f"B{summary_row + 4}"] = f'=IFERROR(AVERAGE(K{start_row}:K{start_row + 11}),"")'
    ws[f"A{summary_row + 5}"] = "Conversion C->P promedio cargada"
    ws[f"B{summary_row + 5}"] = f'=IFERROR(AVERAGE(I{start_row}:I{start_row + 11}),"")'
    ws[f"B{summary_row + 4}"].number_format = "0.00%"
    ws[f"B{summary_row + 5}"].number_format = "0.00%"
    ws[f"B{summary_row + 3}"].number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 11
    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["N"].width = 12
    ws.column_dimensions["O"].width = 48

    ws.freeze_panes = "A6"


def _build_monthly_log_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Registro")

    ws["A1"] = "Registro mensual de decisiones"
    ws["A1"].font = Font(bold=True, size=13)

    headers = [
        "Mes",
        "Que funciono",
        "Que no funciono",
        "Bloqueadores",
        "Decision producto",
        "Decision comercial",
        "Decision soporte",
        "Riesgo principal",
        "Mitigacion",
        "Responsable",
        "Fecha revision",
    ]

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _set_border(cell)

    for row in range(4, 16):
        ws.cell(row=row, column=1, value=f"M{row - 3}")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            _set_border(cell)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 8,
        "B": 28,
        "C": 28,
        "D": 22,
        "E": 24,
        "F": 24,
        "G": 24,
        "H": 20,
        "I": 20,
        "J": 14,
        "K": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"


def generate_workbook() -> None:
    wb = Workbook()
    _build_instructions_sheet(wb)
    _build_dashboard_sheet(wb)
    _build_monthly_log_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    generate_workbook()
    print(f"Workbook generated: {OUTPUT_PATH}")
